from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import openpyxl
import time


def waitUntilLoadIsInvis(seconds):
    WebDriverWait(browser, seconds).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.loadingAnimation__text")))
    browser.execute_script('window.stop();')


def check_exists_selector(selector):
    try:
        browser.find_element_by_css_selector(selector)
    except NoSuchElementException or StaleElementReferenceException:
        return False
    return True


def check_exists_xpath(XPath):
    try:
        browser.find_element_by_xpath(XPath)
    except NoSuchElementException:
        return False
    return True


def expand(seconds):
    while check_exists_selector('.event__more'):
        print(check_exists_selector('.event__more'))
        waitUntilLoadIsInvis(30)
        browser.execute_script("return arguments[0].scrollIntoView(true);", more_matches)
        more_matches.click()
        time.sleep(seconds)


def how_much_detail_stats(flag, req, home, away):
    if flag == 0:
        return major_stats_func(req)
    elif flag == 1:
        return major_stats_func(req), major_stats_time1_func(), major_stats_time2_func(home, away)
    elif flag == 2:
        return major_stats_func(req), minor_stats_and_click_button_func()
    else:
        return major_stats_func(req), major_stats_time1_func(), major_stats_time2_func(home, away), minor_stats_and_click_button_func()


def out_of_field_cards_check(bool_flag):
    if bool_flag:
        statistics.append('НЕ НА ПОЛЕ!!!')


def major_stats_func(req):
    for i in xpath_list_ending:
        request = req + i
        if check_exists_xpath(request):
            event_numbers_all = len(browser.find_elements_by_xpath(request))
            statistics.append(event_numbers_all)
        else:
            statistics.append(0)
    # print(statistics)


def major_stats_time1_func():
    for i in xpath_list_ending[:5]:
        request = xpath_anygame_time1_home + i
        if check_exists_xpath(request):
            event_numbers_all = len(browser.find_elements_by_xpath(request))
            statistics.append(event_numbers_all)
        else:
            statistics.append(0)
    for i in xpath_list_ending[:5]:
        request = xpath_anygame_time1_away + i
        if check_exists_xpath(request):
            event_numbers_all = len(browser.find_elements_by_xpath(request))
            statistics.append(event_numbers_all)
        else:
            statistics.append(0)
    # print(statistics)


def major_stats_time2_func(home, away):
    for i in xpath_list_ending[:5]:
        request = home + i
        if check_exists_xpath(request):
            event_numbers_all = len(browser.find_elements_by_xpath(request))
            statistics.append(event_numbers_all)
        else:
            statistics.append(0)
    for i in xpath_list_ending[:5]:
        request = away + i
        if check_exists_xpath(request):
            event_numbers_all = len(browser.find_elements_by_xpath(request))
            statistics.append(event_numbers_all)
        else:
            statistics.append(0)
    # print(statistics)


def collect_events_from_browse():
    waitUntilLoadIsInvis(5)
    time.sleep(1)
    date = browser.find_element_by_xpath('//div[@id="utime"]').text
    team1 = browser.find_element_by_xpath('//div[@class="team-text tname-home"]//a[@class="participant-imglink"]').text
    team2 = browser.find_element_by_xpath('//div[@class="team-text tname-away"]//a[@class="participant-imglink"]').text
    '''if check_exists_xpath('//div[@class="match-information-data"]/div[1]') and browser.find_element_by_xpath('//div[@class="match-information-data"]/div[1]').text[7] != ':':
        referee = browser.find_element_by_xpath('//div[@class="match-information-data"]/div[1]').text[7:-9]'''
    if check_exists_xpath('//div[@class="match-information-data"]/div[1]'):
        referee = browser.find_element_by_xpath('//div[@class="match-information-data"]/div[1]').text[7:]
    else:
        referee = ''
    date_team1_team2_list = [date, team1, team2, referee]
    statistics.extend(date_team1_team2_list)
    # print(key_number, end=' ')
    if check_exists_selector('.stage-6'):
        req = xpath_cups_extratime
        request_time2_home = xpath_cups_extratime_time2_home
        request_time2_away = xpath_cups_extratime_time2_away
        flag_out_of_field = check_exists_xpath(xpath_cups_extratime + out_of_field)
    elif check_exists_selector('.stage-7'):
        req = xpath_cups_no_extratime
        request_time2_home = xpath_cups_no_extratime_time2_home
        request_time2_away = xpath_cups_no_extratime_time2_away
        flag_out_of_field = check_exists_xpath(out_of_field)
    else:
        req = ''
        request_time2_home = xpath_anygame_time2_home
        request_time2_away = xpath_anygame_time2_away
        flag_out_of_field = check_exists_xpath(out_of_field)
    how_much_detail_stats(how_much_detail_flag, req, request_time2_home, request_time2_away)
    out_of_field_cards_check(flag_out_of_field)


def minor_stats_and_click_button_func():
    browser.find_element_by_xpath('//a[@id="a-match-statistics"]').click()
    time.sleep(0.5)
    waitUntilLoadIsInvis(30)
    stats_elem_list = browser.find_elements_by_xpath('//div[@id="tab-statistics-0-statistic"]//div[@class="statTextGroup"]/*')
    for j in range(len(events)):
        for i in range(1, len(stats_elem_list), 3):
            if events[j] == stats_elem_list[i].text:
                statistics.append(stats_elem_list[i - 1].text)
                statistics.append(stats_elem_list[i + 1].text)


def calc_without_creating_excel():
    red_cards = 0
    penalties = 0
    length = len(data_dict)
    for key, val in data_dict.items():
        if val[5] or val[6] > 0:
            red_cards += 1
        if val[7] or val[8] > 0:
            penalties += 1
        if 'НЕ НА ПОЛЕ!!!' == val[-1]:
            print(val[:3])
    print(red_cards, end=' ')
    print(penalties, end=' ')
    print(length)
    print('Мин кэф на кк нет = ' + str(1 / (1 - (red_cards / length))))
    print('Мин кэф на пень нет = ' + str(1 / (1 - (penalties / length))))


t1 = time.time()
id_list, statistics, temp_list = [], [], []
data_dict = {}
events = ['Угловые', 'Фолы', 'Удары', 'Удары в створ', 'Офсайды', 'Владение мячом', 'Штрафные', 'Всего передач']
# селекторы для всех игр без доп времени и серии пенальти / концевики для остальных селекторов:
y_card = '//span[@class="icon y-card"]'                      # желтая карта
yr_card = '//span[@class="icon yr-card"]'                    # 2-я желтая карта
r_card = '//span[@class="icon r-card"]'                      # красная карта
penalty = '//span[text()="(Пенальти)"]'                      # пенальти
penalty_m = '//span[@class="icon penalty-missed"]'           # незабитый пенальти
substitution = '//span[@class="icon substitution-in"]'       # замена
own_goal = '//span[@class="icon soccer-ball-own"]'           # автогол
xpath_list_ending = [y_card, yr_card, r_card, penalty, penalty_m, substitution, own_goal]
out_of_field = '//span[contains(text(), "Не на поле")]'

# селекторы в кубковых играх (исключаются события в доп время и серия пенальти), доп время есть:
xpath_cups_extratime = '//div[@class="detailMS__incidentsHeader stage-6"]/preceding-sibling::*'

# селекторы в кубковых играх (исключается серия пенальти), доп времени нет:
xpath_cups_no_extratime = '//div[@class="detailMS__incidentsHeader stage-7"]/preceding-sibling::*'

# селекторы для 1-го тайма в любых играх для команды хозяев:
xpath_anygame_time1_home = '//div[contains(@class, "stage-12")]/following-sibling::div[contains(@class, "incidentRow--home") and following::div[contains(@class, "stage-13")]]'

# селекторы для 1-го тайма в любых играх для команды гостей:
xpath_anygame_time1_away = '//div[contains(@class, "stage-12")]/following-sibling::div[contains(@class, "incidentRow--away") and following::div[contains(@class, "stage-13")]]'

# селекторы для 2-го тайма для команды хозяев в кубковых играх (исключаются события в доп время и серия пенальти), доп время есть:
xpath_cups_extratime_time2_home = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--home") and following::div[contains(@class, "stage-6")]]'

# селекторы для 2-го тайма для команды гостей в кубковых играх (исключаются события в доп время и серия пенальти), доп время есть:
xpath_cups_extratime_time2_away = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--away") and following::div[contains(@class, "stage-6")]]'

# селекторы для 2-го тайма для команды хозяев в кубковых играх (исключается серия пенальти), доп времени нет:
xpath_cups_no_extratime_time2_home = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--home") and following::div[contains(@class, "stage-7")]]'

# селекторы для 2-го тайма для команды гостей в кубковых играх (исключается серия пенальти), доп времени нет:
xpath_cups_no_extratime_time2_away = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--home") and following::div[contains(@class, "stage-7")]]'

# селекторы для 2-го тайма в чемп-ах и кубках без доп времени и серии пенальти для команды хозяев:
xpath_anygame_time2_home = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--home")]'

# селекторы для 2-го тайма в чемп-ах и кубках без доп времени и серии пенальти для команды гостей:
xpath_anygame_time2_away = '//div[contains(@class, "stage-13")]/following-sibling::div[contains(@class, "incidentRow--away")]'

# --------------------------------------------------------------------------------------------------------------------------------------
link = 'https://www.flashscore.com.ua/football/england/premier-league/results/'
how_much_detail_flag = 0
# 0 - ж, кж, к, п, нп, з, а
# 1 - 0+ ж, кж, к по таймам и по командам
# 2 - 0+ угл, ф, у, увс, оф, вм, ш, вп
# any - 0+1+2
# --------------------------------------------------------------------------------------------------------------------------------------

options = webdriver.ChromeOptions()
options.add_argument('headless')
caps = DesiredCapabilities.CHROME
browser = webdriver.Chrome(chrome_options=options, desired_capabilities=caps)
caps["pageLoadStrategy"] = "none"
browser.implicitly_wait(0.5)
browser.get(link)
waitUntilLoadIsInvis(30)
if check_exists_selector('.event__more'):
    more_matches = browser.find_element_by_css_selector('.event__more')
expand(3)
# =====================================================================================================================================
all_matches_elems = browser.find_elements_by_css_selector('.event__match')
for elem in all_matches_elems:
    id_list.append(elem.get_attribute('id')[4:])   # для чемпионатов
# =====================================================================================================================================
'''all_matches_elems = browser.find_elements_by_xpath('//div[contains(@class, "event__header ")][1]/following-sibling::div['
                                                   'following-sibling::div[contains(text(), "1/32")]]')
for elem in all_matches_elems:
    if elem.get_attribute('id') != '':
        id_list.append(elem.get_attribute('id')[4:])'''  # для кубков до 1/64 или 1/32 и т.д.
# =====================================================================================================================================
# all_matches_elems = browser.find_elements_by_xpath('//div[contains(@class, "event__header ")][1]/following-sibling::div['
#                                                  'following-sibling::div[contains(text(), "Тур 6")]]')
'''all_matches_elems = browser.find_elements_by_xpath('//div[contains(@class, "event__header ")][1]/following-sibling::div[following-sibling::div[contains(text(), "Финал")]]')
for elem in all_matches_elems:
    if elem.get_attribute('id') != '':
        id_list.append(elem.get_attribute('id')[4:])'''     # для ЛЧ и ЛЕ в стадии ПО
# =====================================================================================================================================
key_number = len(id_list)
try:
    for id in id_list:
        browser.get('https://www.myscore.ru/match/'+id+'/#match-summary')
        # print('https://www.myscore.ru/match/'+id+'/#match-summary')
        waitUntilLoadIsInvis(30)
        time.sleep(1)
        collect_events_from_browse()
        data_dict.update({key_number: statistics.copy()})
        statistics.clear()
        print(key_number, end=' ')
        print(data_dict[key_number])
        key_number -= 1
finally:
    browser.quit()

workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
for key, values in data_dict.items():
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1
# workbook.save(filename=f"{link[-18:-9]}.xlsx")
# workbook.save(filename="eng.xlsx")
link_split = link.split("/")
file_name = f'{link_split[-4]}_{link_split[-3]}.xlsx'
workbook.save(filename=file_name)


calc_without_creating_excel()


t2 = time.time()
print((t2-t1)/60)
print(file_name)




